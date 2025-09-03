import openpyxl, re

class CheckExtractor:
    def __init__(self):
        self.checks_l = []
        self.checks_values_d = {}
        self.not_unique_key_l = []
        self.__processed_keys_l = []
        self.__regkey_pattern = re.compile(r'^(HKLM|HKU|HKEY_LOCAL_MACHINE|HKEY_USERS)\\.*', re.MULTILINE)
        self.__ui_path_pattern = re.compile(r'^(.*\\.*)+$')

    def extract_checks_from_xlsx(self, benchmark_xlsx):
        benchmark_sheet=0 # first sheet
        audit_column=9 # column (I)
        remediation_column=10 # column (J)
        
        # Load the workbook
        workbook = openpyxl.load_workbook(benchmark_xlsx, read_only=True)

        # Select the benchmark sheet
        sheet = workbook.worksheets[benchmark_sheet]

        row=5 # Start at second row to account for headers
        cell_value = "Random to get in loop" # Should not be empty
        # Loops until it finds an empty cell
        while cell_value :
            # Get the cell value at audit column
            cell_value = sheet.cell(row=row, column=audit_column).value
            # Check if string (empty cell are NoneType)
            if isinstance(cell_value, str):
                regkeys_l = self.__get_regkeys_in_audit_cell(cell_value)
                if not regkeys_l:
                    # Get the cell value at remediation column
                    cell_value = sheet.cell(row=row, column=remediation_column).value
                    ui_paths_l = self.__get_ui_paths_in_remediation_cell(cell_value)
                    self.checks_l.append(ui_paths_l)
                    self.__get_values_in_remediation_cell(cell_value, ui_paths_l)
                else:
                    self.checks_l.append(regkeys_l)
                    self.__get_values_in_audit_cell(cell_value, regkeys_l)
            row+=1
        workbook.close()

    def __get_regkeys_in_audit_cell(self, cell_value):
        regkeys_l = []

        cell_value_lines = cell_value.split('\n')
        no_match = True
        # Get all params
        for line in cell_value_lines:
            if self.__regkey_pattern.match(line):
                regkey = re.search(r'^(?:HKLM|HKU|HKEY_LOCAL_MACHINE|HKEY_USERS)\\.*', line).group()
                regkeys_l.append(regkey)
                # Check duplicated key
                self.__check_duplicate_key(regkey)
                no_match = False
        # Check if no match found
        if no_match:
            return None
        # Return reg keys
        return regkeys_l

    def __get_ui_paths_in_remediation_cell(self, cell_value):
        ui_paths_l = []

        cell_value_lines = cell_value.split('\n')
        no_match = True
        for line in cell_value_lines:
            if self.__ui_path_pattern.match(line):
                ui_path = line.split(':', 1)[-1].strip()
                ui_paths_l.append(ui_path)
                # Check duplicated key
                self.__check_duplicate_key(ui_path)
                no_match = False
                break
        # Check if no match found
        if no_match:
            return None
        # Return checks
        return ui_paths_l
    
    def __get_values_in_audit_cell(self, cell_value, regkeys_l):
        cell_value_lines = cell_value.split('\n')
        line_values = ""
        
        for line in cell_value_lines:
            if "REG_" in line:
                
                # Extract registry type and value expression
                reg_info = self.__parse_registry_line(line)
                if reg_info and reg_info['parsed_value']:
                    parsed_value = reg_info["parsed_value"]
                    
                    # Handle key mappings
                    if isinstance(parsed_value, dict) and parsed_value.get('operator') == 'key_mapping':
                        mappings = parsed_value['mappings']
                        
                        for regkey in regkeys_l:
                            # Extract the key name from the registry path (last part after :)
                            key_name = regkey.split(':')[-1] if ':' in regkey else regkey
                            
                            if key_name in mappings:
                                specific_value = mappings[key_name]
                                specific_reg_info = {
                                    'type': reg_info['type'],
                                    'raw_value': str(specific_value),
                                    'condition_type': 'specific_value',
                                    'parsed_value': {'operator': '==', 'value': specific_value}
                                }
                                self.checks_values_d[regkey] = specific_reg_info
                    else:
                        # Handle normal case (same value for all keys)
                        for regkey in regkeys_l:
                            self.checks_values_d[regkey] = reg_info
            else:
                for regkey in regkeys_l:
                    self.checks_values_d[regkey] = None

    def __parse_registry_line(self, line):
        """
        Parse a line containing REG_TYPE and extract structured info
        Returns dict with: type, raw_value, parsed_value, condition_type
        """
        # Find the first REG_TYPE in the line
        reg_type_match = re.search(r'(REG_\w+)', line)
        if not reg_type_match:
            return None
        
        reg_type = reg_type_match.group(1)
        
        # Extract everything after the REG_TYPE
        remainder = line[reg_type_match.end():].strip()
        
        # Different patterns to match
        patterns = [
            # REG_TYPE value of X
            (r'value\s+of\s+(.+)', 'value_of'),
            # REG_TYPE value between X and Y  
            (r'value\s+between\s+(.+)', 'range'),
            # REG_TYPE : (empty or standalone)
            (r':\s*$', 'empty'),
            # REG_TYPE : X (colon followed by value)
            (r':\s*(.+)', 'colon_value'),
        ]
        
        for pattern, condition_type in patterns:
            match = re.search(pattern, remainder, re.IGNORECASE)
            if match:
                raw_value = match.group(1).strip() if match.lastindex else ""
                
                return {
                    'type': reg_type,
                    'raw_value': raw_value,
                    'condition_type': condition_type,
                    'parsed_value': self.__parse_value_expression(raw_value, condition_type)
                }
        
        # If no pattern matched, return basic info
        return {
            'type': reg_type,
            'raw_value': remainder,
            'condition_type': 'unknown',
            'parsed_value': remainder
        }

    def __parse_value_expression(self, raw_value, condition_type):
        """
        Parse the actual value expression into structured data
        """
        if condition_type == 'empty':
            return None
        
        if condition_type == 'range':
            # Handle "between X and Y"
            range_match = re.search(r'(\d+)\s+and\s+(\d+)', raw_value)
            if range_match:
                options = range(int(range_match.group(1)), int(range_match.group(2)) + 1)
                return {'operator': 'in', 'values': options}
        
        if condition_type in ['value_of', 'colon_value']:
            raw_value = raw_value.strip('.,')  # Remove trailing punctuation
            
            # Check for "X for each rule" pattern
            for_each_match = re.search(r'^(\d+|\w+|"[^"]*")\s+for\s+each\s+rule', raw_value, re.IGNORECASE)
            if for_each_match:
                value = for_each_match.group(1).strip('"\'')
                if value.isdigit():
                    return {'operator': '==', 'value': int(value)}
                else:
                    return {'operator': '==', 'value': value}
            
            # Check for "X or less, but not Y" pattern
            but_not_match = re.search(r'(\d+)\s+or\s+less,\s+but\s+not\s+(\d+)', raw_value, re.IGNORECASE)
            if but_not_match:
                max_val = int(but_not_match.group(1))
                excluded_val = int(but_not_match.group(2))
                return {
                    'operator': 'compound',
                    'conditions': [
                        {'operator': '<=', 'value': max_val},
                        {'operator': '!=', 'value': excluded_val}
                    ]
                }
            
            # Check for "X or more, but not Y" pattern (in case it exists)
            but_not_more_match = re.search(r'(\d+)\s+or\s+more,\s+but\s+not\s+(\d+)', raw_value, re.IGNORECASE)
            if but_not_more_match:
                min_val = int(but_not_more_match.group(1))
                excluded_val = int(but_not_more_match.group(2))
                return {
                    'operator': 'compound',
                    'conditions': [
                        {'operator': '>=', 'value': min_val},
                        {'operator': '!=', 'value': excluded_val}
                    ]
                }
            
            # Check for parenthetical key mappings: "0 (KeyName) and 1 (KeyName2)"
            paren_pattern = r'(\d+)\s*\(([^)]+)\)'
            paren_matches = re.findall(paren_pattern, raw_value)
            if paren_matches and ' and ' in raw_value:
                # Create mapping of key names to values
                key_value_map = {}
                for value, key_name in paren_matches:
                    key_value_map[key_name.strip()] = int(value) if value.isdigit() else value
                return {'operator': 'key_mapping', 'mappings': key_value_map}
            
            # Check for comparatives
            if 'or less' in raw_value:
                num_match = re.search(r'(\d+)\s+or\s+less', raw_value)
                if num_match:
                    return {'operator': '<=', 'value': int(num_match.group(1))}
            
            if 'or more' in raw_value:
                num_match = re.search(r'(\d+)\s+or\s+more', raw_value)
                if num_match:
                    return {'operator': '>=', 'value': int(num_match.group(1))}
            
            if 'or greater' in raw_value:
                num_match = re.search(r'(\d+)\s+or\s+greater', raw_value)
                if num_match:
                    return {'operator': '>=', 'value': int(num_match.group(1))}
        
            if 'or fewer' in raw_value:
                num_match = re.search(r'(\d+)\s+or\s+fewer', raw_value)
                if num_match:
                    return {'operator': '<=', 'value': int(num_match.group(1))}
            
            if 'greater than' in raw_value:
                num_match = re.search(r'greater\s+than\s+(\d+)', raw_value)
                if num_match:
                    return {'operator': '>', 'value': int(num_match.group(1))}
            
            # Comma-separated list (with optional "and" at the end)
            if ',' in raw_value and ' or ' not in raw_value and not re.search(r'\([^)]+\)', raw_value):
                # normalize "and" → comma
                normalized = re.sub(r',\s+and\s+', ', ', raw_value, flags=re.IGNORECASE)
                # split and strip everything
                items = [re.sub(r'\s+', '', p.strip().strip('"\''))
                        for p in normalized.split(',')]
                items = [i for i in items if i]  # remove empties
                return {'operator': '==', 'value': items}
            
            # Check for "X or Y" (multiple options)
            if ' or ' in raw_value and all(x not in raw_value for x in ['or less', 'or more', 'or greater', 'or fewer']):
                # Split on 'or'
                parts = raw_value.split(' or ')
                options = []
                for part in parts:
                    # Further split on ',' and clean
                    subparts = [p.strip().strip('"\'') for p in part.split(',')]
                    options.extend([sp for sp in subparts if sp])  # filter empty
                return {'operator': 'in', 'values': options}
            
            # Check for simple numeric value
            if raw_value.isdigit():
                return {'operator': '==', 'value': int(raw_value)}
            
            # Check for quoted string
            quoted_match = re.search(r'["\']([^"\']+)["\']', raw_value)
            if quoted_match:
                return {'operator': '==', 'value': quoted_match.group(1)}
            
            # Return as-is for complex expressions
            return {'operator': '==', 'value': raw_value}
        
        return raw_value

    def __get_values_in_remediation_cell(self, cell_value, ui_paths_l):
        cell_value_lines = cell_value.split('\n')
        line_values = ""
        
        # Get value line
        for line in cell_value_lines:
            if 'UI path to' in line:
                line_values = line.split('UI path to', 1)[-1].split(':', 1)[0]
                break
        value = line_values.strip()
        # Add UI path and value to dict
        self.checks_values_d[ui_paths_l[0]] = value

    def __check_duplicate_key(self, regkey):
        # Regkey
        if regkey.startswith(('HKLM', 'HKU', 'HKEY_LOCAL_MACHINE', 'HKEY_USERS')):
            key = regkey.split(':', 1)[-1]
        # UI path
        else:
            # Extract last part of the UI path
            if ":" in regkey:
                key = regkey.split(":")[-1].strip()
            else:
                key = regkey.split("\\")[-1].strip()
        # Update list of keys that are not unique
        if key in self.__processed_keys_l:
            if key not in self.not_unique_key_l:
                self.not_unique_key_l.append(key)
        else:
            self.__processed_keys_l.append(key)