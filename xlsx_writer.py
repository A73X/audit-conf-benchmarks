import openpyxl

class XlsxWriter:
    def __init__(self):
        self.name = "XlsxWriter"
    
    def set_benchmark_xlsx_path(self, benchmark_xlsx):
        self.benchmark_xlsx = benchmark_xlsx
    
    def write(self, checks, values, proofs, compliances, reasons):
        benchmark_sheet = 0 # first sheet
        compliance_column = 1 # column (A)
        value_column = 14 # column (N)
        proof_column = 15 # column (O)
        reason_column = 16 # column (P)
        row = 5 # Start at second row to account for headers

        # Load the workbook
        workbook = openpyxl.load_workbook(self.benchmark_xlsx)
        # Select the benchmark sheet explicitly
        sheet = workbook.worksheets[benchmark_sheet]

        # Iterate through checks
        for i in range(len(checks)):
            compliance = compliances[i]
            regkeys = checks[i]
            sheet.cell(row=row, column=compliance_column).value = compliance
            sheet.cell(row=row, column=value_column).value = self.__format_extracted(regkeys, values)
            sheet.cell(row=row, column=proof_column).value = self.__format_extracted(regkeys, proofs)
            sheet.cell(row=row, column=reason_column).value = "\n".join(reasons[i])
            row += 1

        # Save the changes
        workbook.save(self.benchmark_xlsx)
        workbook.close()
    
    def __format_extracted(self, regkeys, values_or_proofs):
        cell_value = ""
        for regkey in regkeys:
            if regkey in values_or_proofs.keys():
                for value_or_proof in values_or_proofs[regkey]:
                    if cell_value: # Format new line
                        cell_value += "\n"
                    cell_value += f"{regkey} : {value_or_proof}"
            else:
                if cell_value: # Format new line
                    cell_value += "\n"
                cell_value += f"{regkey} : NOT FOUND"
        return cell_value