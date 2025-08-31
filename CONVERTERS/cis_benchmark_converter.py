#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Class based on "cis_benchmark_converter.py" file from https://github.com/Octomany/cisbenchmarkconverter
File : cis_benchmark_converter.py
Author : Maxime Beauchamp
LinkedIn : https://www.linkedin.com/in/maxbeauchamp/ 
Created : 2024-11-06
"""

import csv
import re
import argparse
import pdfplumber
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule
from openpyxl import Workbook
import os
from colorama import Fore, Style, init

class CisBenchmarkConverter:
    def __init__(self):
        # Regular expressions for extracting recommendations and cleaning text
        self.recommendation_pattern = re.compile(r'^\s*(\d+(?:\.\d+)+)\s+(.+)')  # Matches numbers like 1.1.1, 2.2.2.2, etc.
        self.remove_pattern = re.compile(r'Page\s\d{1,3}|•')
        self.title_pattern = re.compile(r'^(\d+\.\d+(?:\.\d+)*)\s*(\(L\d+\))?\s*(.*)')

        # Pattern to remove page numbers (e.g., "Page 123")
        self.page_number_pattern = re.compile(r'\bPage\s+\d+\b', re.IGNORECASE)

        # Sections to extract
        self.sections = [
            'Profile Applicability:',
            'Description:',
            'Rationale:',
            'Impact:',
            'Audit:',
            'Remediation:',
            'Default Value:',
            'References:',
            'Additional Information:'
        ]

    def remove_page_numbers(self, text):
        return self.page_number_pattern.sub('', text)

    def content_to_string(self, content):
        content_string = ""
        index = 0
        
        while index < len(content):
            line = content[index]
            if line.endswith(('.', '!', '?')):
                content_string += line + '\n'
            elif line.startswith(('HKLM', 'HKU', 'HKEY_LOCAL_MACHINE', 'HKEY_USERS')):
                # Try to repair audit registry key path
                if index < len(content) - 1:  # Check if there is a next line
                    # Check if the next line is another registry key
                    if content[index + 1].startswith(('HKLM', 'HKU', 'HKEY_LOCAL_MACHINE', 'HKEY_USERS')):
                        content_string += '\n' + line
                    else:
                        # Subkey splitted
                        if '\\' in content[index + 1]:
                            content_string += '\n' + line + ' ' + content[index + 1]
                        # Key splitted
                        else:
                            content_string += '\n' + line + content[index + 1]
                        index += 1  # Skip the next line since we've processed it
                else:
                    content_string += '\n' + line
            else:
                content_string += line + ' '
            
            index += 1  # Move to next line
        
        return content_string

    def extract_title_and_version(self, input_file):
        with pdfplumber.open(input_file) as pdf:
            first_page = pdf.pages[0]
            page_text = first_page.extract_text().splitlines()
        title_lines = []
        version = None
        for line in page_text:
            if line.lower().startswith("v") and "-" in line:
                version = line.strip()
                break
            else:
                title_lines.append(line.strip())
        title = " ".join(title_lines) if title_lines else "CIS Benchmark Document"
        return title, version

    # Generate a unique filename if the file already exists
    def generate_unique_filename(self, base_name, extension):
        counter = 1
        file_name = f"{base_name}.{extension}"
        while os.path.exists(file_name):
            file_name = f"{base_name}({counter}).{extension}"
            counter += 1
        return file_name

    def write_output(self, recommendations, output_file, output_format, title, version):
        self.log_info(f"Writing output to {output_file} in {output_format.upper()} format...")

        if output_format == 'csv':
            headers = ['Compliance Status', 'Number', 'Level', 'Title'] + [sec[:-1] for sec in self.sections if sec != 'CIS Controls:'] + ['Machine Value', 'Proofs']
            with open(output_file, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='|')
                writer.writerow([title if title else "CIS Benchmark Document"])
                writer.writerow([version if version else ""])
                writer.writerow([])  # Empty row for spacing
                writer.writerow(headers)  # Column headers

                for recommendation in recommendations:
                    recommendation['Compliance Status'] = 'To Review'
                    row = [recommendation.get(header, '') for header in headers]
                    writer.writerow(row)

        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Recommendations"
            sheet["A1"] = title if title else "CIS Benchmark Document"
            sheet["A1"].font = Font(size=14, bold=True)
            sheet["A2"] = version if version else ""
            sheet["A2"].font = Font(size=12, italic=True)

            headers = ['Compliance Status', 'Number', 'Level', 'Title'] + [sec[:-1] for sec in self.sections if sec != 'CIS Controls:'] + ['Machine Value', 'Proofs']
            sheet.append([""] * len(headers))  # Empty row for spacing
            sheet.append(headers)

            for row_idx, recommendation in enumerate(recommendations, start=5):
                recommendation['Compliance Status'] = 'To Review'
                row = [recommendation.get(header, '') for header in headers]
                sheet.append(row)

            dv = DataValidation(type="list", formula1='"Compliant,Non-Compliant,To Review"', showDropDown=False)
            sheet.add_data_validation(dv)
            for row_idx in range(5, len(recommendations) + 5):
                dv.add(sheet[f"A{row_idx}"])

            compliant_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            non_compliant_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            to_review_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            compliant_rule = FormulaRule(formula=['$A5="Compliant"'], fill=compliant_fill)
            non_compliant_rule = FormulaRule(formula=['$A5="Non-Compliant"'], fill=non_compliant_fill)
            to_review_rule = FormulaRule(formula=['$A5="To Review"'], fill=to_review_fill)
            
            sheet.conditional_formatting.add(f"A5:A{len(recommendations) + 5}", compliant_rule)
            sheet.conditional_formatting.add(f"A5:A{len(recommendations) + 5}", non_compliant_rule)
            sheet.conditional_formatting.add(f"A5:A{len(recommendations) + 5}", to_review_rule)

            # Add table style
            tab = Table(displayName="CISRecommendations", ref=f"A4:{chr(65+len(headers)-1)}{len(recommendations) + 4}")
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            sheet.add_table(tab)

            # Set column widths
            sheet.column_dimensions['A'].width = 10  # Compliance Status
            sheet.column_dimensions['B'].width = 8  # Number (default width)
            sheet.column_dimensions['C'].width = 8  # Level (default width)
            sheet.column_dimensions['D'].width = 50  # Title
            for col in range(5, 13):  # Columns E to L (Profile Applicability to References)
                sheet.column_dimensions[chr(64 + col)].width = 10

            workbook.save(output_file)
            workbook.close()

        self.log_info(f"Finished writing {len(recommendations)} recommendations to {output_file}.")

    # Logging functions
    def log_info(self, message):
        print(f"\n{Fore.GREEN}[INFO]{Style.RESET_ALL} {message}")

    def log_warning(self, message):
        print(f"\n{Fore.YELLOW}[WARNING]{Style.RESET_ALL} {message}")

    def log_debug(self, message):
        print(f"\n{Fore.BLUE}[DEBUG]{Style.RESET_ALL} {message}")

    def read_pdf(self, input_file):
        self.log_info("Starting to read the PDF file...")
        text = []
        with pdfplumber.open(input_file) as pdf:
            total_pages = len(pdf.pages)
            extraction_started = False
            
            # Start reading from page 10 to skip the table of contents
            for page_number, page in enumerate(pdf.pages[9:], start=10):
                page_text = page.extract_text()
                
                # Display progress
                print(f"\r{Fore.GREEN}[INFO]{Style.RESET_ALL} Processing page {page_number}/{total_pages}...", end="", flush=True)
                
                if not extraction_started:
                    if "Recommendations" in page_text and "....." not in page_text and "Recommendation Definitions" not in page_text:
                        extraction_started = True
                        self.log_debug(f"Recommendations section detected. Starting extraction... (This may take a while)")
                
                if extraction_started:
                    if "Appendix: Summary Table" in page_text or "Checklist" in page_text:
                        self.log_debug("End of Recommendations section reached.")
                        break
                    text.append(page_text)

        self.log_info("\nCompleted reading the PDF file.")
        return '\n'.join(text)

    def find_profile_applicability(self, lines, start_index, max_depth=10):
        """
        Look for 'Profile Applicability:' within a certain depth from the start index.
        Returns True if found within the limit, otherwise False.
        """
        for i in range(start_index + 1, min(start_index + max_depth, len(lines))):
            line = lines[i].strip()
            
            # Check for "Profile Applicability:"
            if line.startswith("Profile Applicability:"):
                return True
            
            # Stop if another title or section is detected
            if self.title_pattern.match(line) or any(line.startswith(sec) for sec in self.sections):
                return False
        
        return False

    def extract_recommendations(self, text):
        """
        Extract recommendations while avoiding duplicates and confirming section content.
        """
        recommendations = []
        lines = text.splitlines()
        current_recommendation = {}
        current_index = 0

        while current_index < len(lines):
            line = lines[current_index].strip()
            line = self.remove_page_numbers(line)  # Remove any page number mentions

            # Utilisation dans le contexte principal
            title_match = self.title_pattern.match(line)
            if title_match:
                # Utilise find_profile_applicability pour vérifier dynamiquement
                if self.find_profile_applicability(lines, current_index):
                    # Sauvegarde la recommandation précédente
                    if current_recommendation:
                        recommendations.append(current_recommendation)
                    
                    # Initialiser une nouvelle recommandation sans doublons
                    current_recommendation = {
                        'Number': title_match.group(1),
                        'Level': title_match.group(2) or '',
                        'Title': title_match.group(3),
                    }
                    
                    # Capture multi-line titles
                    while (
                        current_index + 1 < len(lines) and
                        not any(lines[current_index + 1].strip().startswith(sec) for sec in self.sections) and
                        not self.title_pattern.match(lines[current_index + 1].strip())
                    ):
                        current_index += 1
                        current_recommendation['Title'] += " " + lines[current_index].strip()

            # Capture sections for the current recommendation
            for section in self.sections:
                if line.startswith(section):
                    content, next_index = self.extract_section(lines, current_index, section)
                    current_recommendation[section[:-1]] = content  # Exclude the colon
                    current_index = next_index - 1  # Adjust index after extraction
                    break
            
            current_index += 1

        # Final recommendation
        if current_recommendation:
            recommendations.append(current_recommendation)
        
        # Remove duplicates based on recommendation number and title
        unique_recommendations = { (rec['Number'], rec['Title']): rec for rec in recommendations }
        return list(unique_recommendations.values())

    def extract_section(self, lines, start_index, section_name):
        """
        Extract content of a section until encountering another section or title.
        Lines containing "CIS Controls" are excluded.
        """
        content = []
        current_index = start_index + 1
        while current_index < len(lines):
            line = lines[current_index].strip()
            line = self.remove_page_numbers(line)  # Clean each line of page numbers

            # Stop at new section, title, or "CIS Controls"
            if any(line.startswith(sec) for sec in self.sections) or self.title_pattern.match(line) or 'CIS Controls' in line:
                break

            content.append(line)
            current_index += 1
        
        return self.content_to_string(content).strip(), current_index
    
    def convert(self, benchmark_pdf, output_format):
        base_name = os.path.splitext(os.path.basename(benchmark_pdf))[0]
        extension = "csv" if output_format == "csv" else "xlsx"
        output_file = self.generate_unique_filename(base_name, extension)
        title, version = self.extract_title_and_version(benchmark_pdf)
        text = self.read_pdf(benchmark_pdf)
        recommendations = self.extract_recommendations(text)
        self.write_output(recommendations, output_file, output_format, title, version)
        return output_file